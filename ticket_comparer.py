import openpyxl
import jellyfish
from datetime import datetime
from cleaner import regex_cleaner


class Classificazione:

    def __init__(self, ticket, cluster, e):
        self.ticket = ticket
        self.cluster = cluster
        self.e = e

    def __str__(self):
        tmp_ticket = self.ticket.copy()
        tmp_ticket["descrizione"] = ""
        return f"\n\n\nclass: {self.e}, valore_ticket: {str(tmp_ticket)}"

    def __repr__(self):
        return str(self)

    def set_ticket_cluster(self, cluster, score):
        self.cluster = cluster
        self.e = score
        return self


class Score:

    def __init__(self, damerau_score, jaccard_score, average_score):
        self.damerau_score = damerau_score
        self.jaccard_score = jaccard_score
        self.average_score = average_score

    def __repr__(self):
        return str(self)


def create_clusters(tickets, threshold, upper_threshold):
    # threshold -> se stiamo parlando di probabilità: p == 0,5 => tutto può essere, p > 0,5 => più probabile, p < 0,5 meno improbabile
    # probabilità è un valore tra 0 e 1
    #

    clusters = {}
    cluster_count = 0

    for i in range(len(tickets)):
        ticket_i = tickets[i]
        cluster = []
        cluster_key = ""
        cluster_count += 1
        avanzamento = 0

        if ticket_i.e is not None and ticket_i.e.average_score >= upper_threshold:
            print(
                f"ATTENZIONE: è stato saltato il cluster {ticket_i.ticket["idReclamo"]}!"
            )
            continue
        # TODO: Riattivare questo pezzo di codice, in cui causa un crash per lista disallineata per bug
        #    cluster_key = ticket_i.cluster
        else:
            cluster_key = f"{ticket_i.ticket["idReclamo"]}"

        print(f"Clustering per ticket {cluster_key}")

        for j in range(i, len(tickets)):
            ticket_j = tickets[j]

            if ticket_j.e is not None and ticket_j.e.average_score >= upper_threshold:
                continue

            score = Score(
                d_damerau_normalizzata(ticket_j.ticket, ticket_i.ticket),
                0,
                0,
            )

            score.average_score = calculate_score(score)

            if score.average_score < threshold:
                continue

            if ticket_j.cluster is None:
                cluster.append(ticket_j)
                tickets[j].set_ticket_cluster(cluster_key, score)
                continue

            if ticket_j.e.average_score < score.average_score:
                prev_cluster_key = tickets[j].cluster
                clusters[prev_cluster_key].remove(ticket_j)
                cluster.append(ticket_j)
                tickets[j].set_ticket_cluster(cluster_key, score)

        clusters[cluster_key] = cluster

        if cluster_count % 20 == 0:
            print(f"Generando excel con {cluster_count} clusters")
            print_excel_workbook(clusters)

    return clusters


def calculate_score(ticket_j, ticket_i):
    # TODO: Implementa un sistema di score weighting
    return d_damerau_normalizzata(ticket_j, ticket_i)


def calculate_score(score):
    return score.damerau_score


def d_jaccard(ticket_j, ticket_i):
    return jellyfish.jaccard_similarity(
        ticket_j["descrizione"], ticket_i["descrizione"]
    )


def d_damerau_normalizzata(ticket_j, ticket_i):
    # -> riferimento al documento
    damerau_distance = jellyfish.damerau_levenshtein_distance(
        ticket_j["descrizionePulita"], ticket_i["descrizionePulita"]
    )

    max_length = max(
        len(ticket_i["descrizionePulita"]),
        len(ticket_j["descrizionePulita"]),
    )

    return (
        (1 - (damerau_distance / max_length))
        if damerau_distance != 0 or max_length != 0
        else 0
    )


def print_excel_workbook(clusters):
    workbook = openpyxl.Workbook()
    cluster_count = 0

    for cluster_key in clusters:
        if len(clusters[cluster_key]) <= 1:
            continue

        cluster_sheet = workbook.create_sheet(title=cluster_key)

        cluster_sheet["A1"] = "idReclamo"
        cluster_sheet["B1"] = "damerau"
        cluster_sheet["C1"] = "score"
        cluster_sheet["D1"] = "descrizione"
        cluster_sheet["E1"] = "descrizione Pulita"

        cell_count = 2

        cluster = clusters[cluster_key]

        for tt in cluster:
            ticket = tt.ticket
            score = tt.e
            cluster_sheet[f"A{cell_count}"] = ticket["idReclamo"]
            cluster_sheet[f"B{cell_count}"] = score.damerau_score
            cluster_sheet[f"C{cell_count}"] = score.average_score
            cluster_sheet[f"D{cell_count}"] = ticket["descrizione"]
            cluster_sheet[f"E{cell_count}"] = ticket["descrizionePulita"]
            cell_count += 1

        cluster_count += 1

    # workbook.remove("sheet")
    workbook.save(create_sheet_filename("test"))

    print(f"Generato foglio excel contenente {cluster_count} clusters")


def create_sheet_filename(
    filename: str, extension: str = "xlsx", directory: str = "zoutput"
):
    return f"{directory}/{filename}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.{extension}"


# def calculate_e2(ticket_j, ticket_i):
# 	a = 10
# 	b = 15
# 	return (a*d_jaccard(ticket_j, ticket_i) + *d_damerau_normalizzata(ticket_j, ticket_i))/(a + b)

# TODO: estrazione automatica dei ticket da giacenza invece di estrazione manuale
file = "giacenza.xlsx"

print("Caricamento della giacenza in corso")
workbook = openpyxl.open(file)
sheet_ticket = workbook.active

colonne = sheet_ticket.max_column
tickets = sheet_ticket.max_row - 1

lista_ticket = []

print("Caricamento dei ticket in corso")
for tt in range(2, tickets + 2):
    ticket_object = {
        "idReclamo": sheet_ticket.cell(tt, 1).value,
        "descrizione": sheet_ticket.cell(tt, 2).value,
        "descrizionePulita": regex_cleaner(sheet_ticket.cell(tt, 2).value.lower()),
    }

    if ticket_object["descrizionePulita"] != "":
        lista_ticket.append(Classificazione(ticket_object, None, None))

print(f"Sono stati caricati {len(lista_ticket)} tickets")

clusters = create_clusters(lista_ticket, 0.53, 0.68)

# print(clusters)

print_excel_workbook(clusters)
